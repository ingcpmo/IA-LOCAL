"""W5 V2 -- generador de documento candidato para XLSX
(CORRECTED_DOCUMENT_GENERATION_AND_FORMAT_SPEC.md, estrategia XLSX de la
sección 14 del plan: "preservar hojas, fórmulas, tablas y rangos;
registrar cambios por hoja y celda -- redline como registro celda a
celda").

Mismo patrón ya usado en `candidate_document_generator.py` (Fase J) para
DOCX: `generate_candidate_workbook` (versión limpia) +
`generate_redline_workbook` (versión marcada + manifest de inserción),
solo que aquí el "documento" es un workbook real y la unidad de cambio es
una celda, no un párrafo dentro de una sección.

Direccionamiento de celda: `RemediationChange.document_location` debe
tener la forma `"NombreHoja!CELDA"` (p.ej. `"Alarms!C12"`) -- formato
determinista, sin ambigüedad, verificado antes de aplicar cualquier
cambio.

Igual que DOCX: solo `CONTENT_ADDITION`/`CONTENT_REPLACEMENT` tienen
sentido a nivel de celda (una celda siempre tiene un valor, "agregar" y
"reemplazar" son la misma operación mecánica -- sobrescribir el valor).
No se restringe por `change_type` aquí a propósito: a diferencia de un
párrafo de texto corrido, una celda no tiene noción de "insertar sin
reemplazar" -- siempre se fija el nuevo valor."""
from __future__ import annotations

import hashlib
import re

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Font
from openpyxl.workbook import Workbook

_CELL_LOCATION_RE = re.compile(r"^([^!]+)!([A-Z]+[0-9]+)$")

_INSERTED_FONT_COLOR = "FF008000"  # verde, mismo criterio visual que DOCX (_INSERTED_COLOR)


def _sha256_text(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def _parse_cell_location(document_location: str) -> tuple[str, str]:
    """document_location -> (sheet_name, cell_coordinate). Lanza
    ValueError explícito si el formato no es 'Hoja!CELDA' -- nunca adivina
    una celda por defecto."""
    match = _CELL_LOCATION_RE.match(document_location.strip())
    if not match:
        raise ValueError(
            f"document_location {document_location!r} no tiene el formato esperado 'NombreHoja!CELDA' "
            "(p.ej. 'Alarms!C12') -- no se adivina una celda por defecto."
        )
    return match.group(1), match.group(2)


def _validate_change_against_workbook(wb: Workbook, change: dict) -> tuple[str, str]:
    sheet_name, cell_coord = _parse_cell_location(change["document_location"])
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"change_id={change['change_id']!r}: hoja '{sheet_name}' no existe en el workbook real "
            f"(hojas reales: {wb.sheetnames})"
        )
    return sheet_name, cell_coord


def generate_candidate_workbook(original_xlsx_path: str, changes: list[dict]) -> Workbook:
    """Versión limpia (§14): abre el XLSX original REAL (nunca se
    modifica en disco -- openpyxl carga en memoria, el archivo fuente
    permanece intacto), preserva TODAS las hojas/fórmulas/rangos, y
    sobrescribe únicamente las celdas declaradas en cada change, sin
    marcado visual."""
    wb = openpyxl.load_workbook(original_xlsx_path, data_only=False)
    for change in changes:
        sheet_name, cell_coord = _validate_change_against_workbook(wb, change)
        wb[sheet_name][cell_coord] = change["proposed_content"]
    return wb


def generate_redline_workbook(
    original_xlsx_path: str, changes: list[dict]
) -> tuple[Workbook, list[dict]]:
    """Redline real (§14): mismo workbook, pero la celda modificada queda
    en verde con un comentario `[change_id]` -- mismo criterio visual que
    el redline DOCX (`_INSERTED_COLOR`). Devuelve (Workbook,
    insertion_manifest): [{change_id, sheet_name, cell_coordinate,
    original_value, proposed_content_sha256}]."""
    wb = openpyxl.load_workbook(original_xlsx_path, data_only=False)
    insertion_manifest: list[dict] = []
    for change in changes:
        sheet_name, cell_coord = _validate_change_against_workbook(wb, change)
        ws = wb[sheet_name]
        cell = ws[cell_coord]
        original_value = cell.value
        cell.value = change["proposed_content"]
        cell.font = Font(color=_INSERTED_FONT_COLOR, bold=True)
        cell.comment = Comment(
            f"[{change['change_id']}] Valor original: {original_value!r}", "AGT-DOC (W5 V2)",
        )
        insertion_manifest.append({
            "change_id": change["change_id"], "sheet_name": sheet_name, "cell_coordinate": cell_coord,
            "original_value": original_value,
            "proposed_content_sha256": _sha256_text(str(change["proposed_content"])),
        })
    return wb, insertion_manifest


def verify_workbook_conformance(
    xlsx_path: str, changes: list[dict], insertion_manifest: list[dict]
) -> list[dict]:
    """`DOCUMENT_CONFORMANCE` para XLSX (mismo principio que
    `candidate_document_generator.verify_document_conformance`): reabre el
    `.xlsx` YA GUARDADO en disco (nunca el objeto Workbook en memoria) y
    verifica que el valor real de cada celda coincide con
    `proposed_content`."""
    reopened = openpyxl.load_workbook(xlsx_path, data_only=False)
    manifest_by_change_id = {m["change_id"]: m for m in insertion_manifest}

    results: list[dict] = []
    for change in changes:
        change_id = change["change_id"]
        entry = manifest_by_change_id.get(change_id)
        if entry is None:
            results.append({
                "change_id": change_id, "status": "CHANGE_NOT_APPLIED",
                "reason": "sin entrada en insertion_manifest",
            })
            continue
        if entry["sheet_name"] not in reopened.sheetnames:
            results.append({
                "change_id": change_id, "status": "CHANGE_NOT_APPLIED",
                "reason": f"hoja '{entry['sheet_name']}' no existe en el xlsx reabierto",
            })
            continue
        actual_value = reopened[entry["sheet_name"]][entry["cell_coordinate"]].value
        expected_value = change["proposed_content"]
        if str(actual_value) == str(expected_value) and _sha256_text(str(actual_value)) == entry["proposed_content_sha256"]:
            results.append({
                "change_id": change_id, "status": "DOCUMENT_CONFORMANCE",
                "sheet_name": entry["sheet_name"], "cell_coordinate": entry["cell_coordinate"],
            })
        else:
            results.append({
                "change_id": change_id, "status": "CHANGE_NOT_APPLIED",
                "reason": "valor de celda no coincide con proposed_content",
            })
    return results
